"""
QLingo - Report Export Module
Handles exporting QA results to various formats (Excel, PDF, CSV)
"""

import pandas as pd
import io
from datetime import datetime
from typing import Dict, Optional

# For PDF export
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("ReportLab not available. Install with: pip install reportlab")

# For Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("openpyxl not available. Install with: pip install openpyxl")


class ReportExporter:
    """Handle exporting QA results to different formats"""
    
    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def export(self, df: pd.DataFrame, summary: Dict, format_type: str) -> Optional[bytes]:
        """
        Export results to specified format
        
        Args:
            df: Results DataFrame
            summary: Summary statistics dictionary
            format_type: Export format ('Excel', 'PDF', or 'CSV')
            
        Returns:
            Bytes of the exported file or None if export fails
        """
        if format_type == "Excel":
            return self.export_to_excel(df, summary)
        elif format_type == "PDF":
            return self.export_to_pdf(df, summary)
        elif format_type == "CSV":
            return self.export_to_csv(df)
        else:
            return None
    
    def export_to_csv(self, df: pd.DataFrame) -> bytes:
        """Export to CSV format"""
        output = io.StringIO()
        df.to_csv(output, index=False)
        return output.getvalue().encode('utf-8')
    
    def export_to_excel(self, df: pd.DataFrame, summary: Dict) -> Optional[bytes]:
        """Export to Excel with formatting"""
        if not EXCEL_AVAILABLE:
            # Fallback to basic Excel export
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='QA Results')
            return output.getvalue()
        
        output = io.BytesIO()
        
        # Create workbook
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write summary sheet
            summary_df = pd.DataFrame([
                ['Report Generated', self.timestamp],
                ['', ''],
                ['Total Segments', summary.get('total_segments', 0)],
                ['Average Quality Score', f"{summary.get('avg_quality', 0):.1f}%"],
                ['Segments with Errors', summary.get('segments_with_errors', 0)],
                ['Total Errors Found', summary.get('total_errors', 0)]
            ], columns=['Metric', 'Value'])
            
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Write results sheet
            df.to_excel(writer, sheet_name='QA Results', index=False)
            
            # Format summary sheet
            workbook = writer.book
            summary_sheet = writer.sheets['Summary']
            
            # Header formatting
            header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Title formatting
            title_font = Font(bold=True, size=14, color='2E86AB')
            for row in [3, 4, 5, 6]:
                summary_sheet[f'A{row}'].font = title_font
            
            # Column widths
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 20
            
            # Format results sheet
            results_sheet = writer.sheets['QA Results']
            
            # Header formatting
            for cell in results_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for idx, col in enumerate(df.columns, 1):
                column_letter = openpyxl.utils.get_column_letter(idx)
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                ) + 2
                results_sheet.column_dimensions[column_letter].width = min(max_length, 50)
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in results_sheet.iter_rows(min_row=1, max_row=len(df)+1):
                for cell in row:
                    cell.border = thin_border
        
        return output.getvalue()
    
    def export_to_pdf(self, df: pd.DataFrame, summary: Dict) -> Optional[bytes]:
        """Export to PDF format"""
        if not PDF_AVAILABLE:
            return None
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2E86AB'),
            spaceAfter=20,
            alignment=1  # Center
        )
        elements.append(Paragraph("QLingo Translation QA Report", title_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Timestamp
        timestamp_style = ParagraphStyle(
            'Timestamp',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        elements.append(Paragraph(f"Generated: {self.timestamp}", timestamp_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Summary section
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2E86AB'),
            spaceAfter=15
        )
        elements.append(Paragraph("Executive Summary", subtitle_style))
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Segments", str(summary.get('total_segments', 0))],
            ["Average Quality Score", f"{summary.get('avg_quality', 0):.1f}%"],
            ["Segments with Errors", str(summary.get('segments_with_errors', 0))],
            ["Total Errors Found", str(summary.get('total_errors', 0))]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.4 * inch))
        
        # Results preview (first 10 rows)
        elements.append(Paragraph("Quality Analysis Results (Preview)", subtitle_style))
        
        # Prepare data for table (limit to first 10 rows)
        preview_df = df.head(10)
        table_data = [preview_df.columns.tolist()] + preview_df.values.tolist()
        
        # Convert all values to strings and truncate long text
        for i in range(len(table_data)):
            for j in range(len(table_data[i])):
                text = str(table_data[i][j])
                if len(text) > 50:
                    text = text[:47] + "..."
                table_data[i][j] = text
        
        # Create table with adjusted column widths
        col_widths = [0.6*inch, 2*inch, 2*inch, 0.8*inch, 0.7*inch, 0.7*inch]
        results_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        results_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            # Body
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        
        elements.append(results_table)
        
        # Footer note
        if len(df) > 10:
            elements.append(Spacer(1, 0.2 * inch))
            note_style = ParagraphStyle(
                'Note',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.grey,
                alignment=1
            )
            elements.append(Paragraph(
                f"Note: Showing first 10 of {len(df)} segments. Download full report for complete results.",
                note_style
            ))
        
        # Build PDF
        doc.build(elements)
        
        return output.getvalue()
    
    @staticmethod
    def get_mime_type(format_type: str) -> str:
        """Get MIME type for format"""
        mime_types = {
            "Excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "PDF": "application/pdf",
            "CSV": "text/csv"
        }
        return mime_types.get(format_type, "application/octet-stream")